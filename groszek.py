import fdb
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font, alignment
from openpyxl.comments import Comment
import yaml
from bs4 import BeautifulSoup
import httpx
import wikitextparser as wtp
import sys


def get_config():
    global config
    try:
        with open(r'config.yml', encoding='utf8') as yf:
            config = yaml.full_load(yf)
    except:
        print('brak pliku config.yml')
        sys.exit(1)


def get_connection(db_name):
    try:
        con = fdb.connect(
            host=config['server'], database=config['dbpath'] + db_name,
            user=config['user'], password=config['password'],
            charset='WIN1250'
        )
        return con
    except:
        print('błąd połaczenia z bazą')
        sys.exit(1)


def get_data(db_name):
    query = '''
    SELECT TRIM(rr.nazwa) || ': ' || TRIM(f.OPIS) || ': ' || f.KOD_FUNSYS uprawnienie, r.NAZWA uzytkownik, TRIM(rr.opis), TRIM(f.opis)
    FROM OP_SLFUN f
    LEFT OUTER JOIN IS_REJESTR rr ON f.KOD_SYSTEMU = rr.ID_SYSTEMU
    LEFT OUTER JOIN OP_OPFUN oof ON f.KOD_FUNSYS = oof.KOD_FUNSYS AND f.KOD_SYSTEMU = oof.KOD_SYSTEMU
    LEFT OUTER JOIN OP_OPER r ON r.KOD_OPER = oof.KOD_OPER AND r.STATUS = 'T'
    UNION
    SELECT TRIM(rr.nazwa) || ': ' || TRIM(f.OPIS) || ': ' || f.KOD_FUNSYS uprawnienie, r.NAZWA uzytkownik, TRIM(rr.opis), TRIM(f.opis)
    FROM OP_SLFUN f
    LEFT OUTER JOIN IS_REJESTR rr ON f.KOD_SYSTEMU = rr.ID_SYSTEMU
    LEFT OUTER JOIN OP_GRFUN oof ON f.KOD_FUNSYS = oof.KOD_FUNSYS AND f.KOD_SYSTEMU = oof.KOD_SYSTEMU
    LEFT OUTER JOIN OP_OPERGRUP og ON og.KOD_GRUPY = oof.KOD_GRUPY
    LEFT OUTER JOIN OP_OPER r ON r.KOD_OPER = og.KOD_OPER AND r.STATUS = 'T'
    '''
    con = get_connection(db_name)
    cur = con.cursor()
    cur.execute(query)
    data = cur.fetchall()
    data_dict = [{'permission': permission, 'username': username, 'system': system, 'description': description}
                 for permission, username, system, description in data]
    cur.close()
    return (data_dict)

def get_groups(db_name):
    query = '''
    SELECT TRIM(rr.nazwa) || ': ' || TRIM(f.OPIS) || ': ' || f.KOD_FUNSYS uprawnienie, TRIM(rr.opis) system, TRIM(g.Nazwa) ggroup
    FROM OP_SLFUN f
    LEFT OUTER JOIN IS_REJESTR rr ON f.KOD_SYSTEMU = rr.ID_SYSTEMU
    LEFT OUTER JOIN OP_GRFUN oof ON f.KOD_FUNSYS = oof.KOD_FUNSYS AND f.KOD_SYSTEMU = oof.KOD_SYSTEMU
    LEFT OUTER JOIN OP_GRUPY g ON g.KOD_GRUPY = oof.KOD_GRUPY
    '''
    con = get_connection(db_name)
    cur = con.cursor()
    cur.execute(query)
    data = cur.fetchall()
    data_dict = [{'permission': permission, 'system': system, 'group': ggroup}
                 for permission, system, ggroup in data]
    cur.close()
    return (data_dict)


def create_worksheet(db_name, wb):
    # pobranie grup uprawineń
    data_groups = get_groups(db_name)
    groups = list({i['group'] for i in data_groups if i['group'] is not None})
    groups.sort(key=lambda s: s[0].lower())
    groups = [{'id': i, 'group': v} for i, v in enumerate(groups)]
    number_of_groups = len(groups)
    # pobranie uprawnień z bazy
    data = get_data(db_name)
    upr = list({(i['permission'], i['system'], i['description']) for i in data})
    upr.sort()
    priviledges = [{'id': i, 'permission': v[0], 'system': v[1], 'description': v[2]}
                   for i, v in enumerate(upr)]
    users = list({i['username'] for i in data if i['username'] is not None})
    users.sort(key=lambda s: s[0].lower())
    users = [{'id': i, 'username': v} for i, v in enumerate(users)]
    # tworzenie arkusza Excela
    ws = wb.create_sheet(title=db_name)
    # kolumna z uprawnieniami
    for priv in priviledges:
        ws[f'A{priv["id"] + 2}'] = priv['permission']
        desc = get_desc(priv['system'], priv['description'])
        if len(desc) and desc != 'Opis':
            ws[f'A{priv["id"] + 2}'].comment = Comment(priv["system"] + ':\n' + desc, '', width=500, height=100)
    # kolumny z grupami uprawnień
    # nagłówek grup
    for g in groups:
        cc = ws.cell(row=1, column=g['id'] + 2)
        cc.value = g['group']
        cc.alignment = Alignment(textRotation=90)
        cc.fill = PatternFill(patternType='solid', fgColor='f0f0f0')
        s = cc.column_letter
        ws.column_dimensions[s].width = 3
    # dane grup
    for g in groups:
        for priv in priviledges:
            if {'permission': priv['permission'], 'group': g['group'],
                'system': priv['system']} in data_groups:
                color = 'b0f0b0'
            else:
                color = 'f0f0f0'
            c = ws.cell(row=priv['id'] + 2, column=g['id'] + 2)
            c.fill = PatternFill(patternType='solid', fgColor=color)
    # użytkownicy
    for user in users:
        cc = ws.cell(row=1, column=user['id'] + 2 + number_of_groups)
        cc.value = user['username']
        cc.alignment = Alignment(textRotation=90)
        s = cc.column_letter
        ws.column_dimensions[s].width = 3
    ws.column_dimensions['A'].width = 105
    for user in users:
        for priv in priviledges:
            if {'permission': priv['permission'], 'username': user['username'],
                'system': priv['system'], 'description': priv['description']} in data:
                c = ws.cell(row=priv['id'] + 2, column=user['id'] + 2 + number_of_groups)
                c.fill = PatternFill(patternType='solid', fgColor='b0b0b0')
    # końcowe formatowanie tabeli - ramki, blokada wiersza
    for column in range(1, len(users) + 2 + number_of_groups):
        for row in range(1, len(upr) + 2):
            ws.cell(row=row, column=column).border = Border(
                left=Side(style='thin'),
                right=Side(style='medium' if column in(1,number_of_groups +1) else 'thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin'))
    ws.freeze_panes = ws['A2']

def get_descriptions():
    global parsed
    page = httpx.get('https://wiki.groszek.pl/index.php?title=Uprawnienia_operator%C3%B3w&action=edit')
    soup = BeautifulSoup(page.text, 'html.parser')
    ta = soup.find('textarea')
    parsed = wtp.parse(ta.string)
    return parsed


def get_desc(system, priv):
    system = system.strip()
    if system in config['systems']:
        sys_wiki = config['systems'][system].lower()
        for section in parsed.sections:
            if section.title:
                sl = (section.title.strip()).lower()
                if sl == sys_wiki:
                    for x in section.tables[0].data():
                        if x[0].lower() == priv.lower():
                            return x[1]
    return ''


if (__name__) == '__main__':
    get_config()
    get_descriptions()
    wb = Workbook()
    wb.remove(wb.active)
    for db_name in config['databases']:
        create_worksheet(db_name, wb)
    wb.save(config.get('excelpath', 'upr_groszek.xlsx'))
